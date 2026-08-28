"""
Employee Hours Roster Updater
=============================
Merges a weekly "Employee Hours" export into the cumulative historical roster.

  1. Sums the new week's hours into the running totals
  2. Appends new hires with their information
  3. Overwrites attribute fields that changed (status, department, supervisor, etc.)
  4. Persists the cumulative historical roster in Google Sheets
  5. Archives each weekly input so prior weeks can be safely replaced/replayed

Matching key: Employee Full Name + Hire Date (strict).
"""

import csv
import hashlib
import io
import re
from datetime import datetime, timezone

import gspread
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Schema
# ----------------------------------------------------------------------------

# Canonical column order of the historical file (Book1), columns A-V.
CANONICAL_COLUMNS = [
    "Employment Status",                        # A
    "Employee Full Name",                       # B
    "Employee Pay Rule",                        # C
    "Temp Agency Code",                         # D
    "Temp Hourly Rates",                        # E
    "Hire Date",                                # F
    "Rehire Date",                              # G
    "Profit Center",                            # H
    "Shift",                                    # I
    "Department",                               # J
    "Reports To",                               # K
    "Regular Hours",                            # L
    "Overtime Hours",                           # M
    "Productive Hours",                         # N
    "Holiday Credit HOLCR",                     # O
    "Absence - Unplanned Hours - Excused",      # P
    "Absence - Planned Hours - Excused",        # Q
    "Absence - PTO Unplanned Hours - Excused",  # R
    "Leave",                                    # S
    "Actual Hours",                             # T
    "Job",                                      # U
    "Absence  - Unplanned Hours - Unexcused",   # V  (note: double space, as in source)
]

# The weekly export renames three absence columns. Mapped positionally.
COLUMN_ALIASES = {
    "Absence Unplanned Hours": "Absence - Unplanned Hours - Excused",
    "Planned Absenteeism Hours": "Absence - Planned Hours - Excused",
    "Extended Absenteeism Hours": "Absence - PTO Unplanned Hours - Excused",
}

# Columns that accumulate week over week.
HOURS_COLUMNS = [
    "Regular Hours",
    "Overtime Hours",
    "Productive Hours",
    "Holiday Credit HOLCR",
    "Absence - Unplanned Hours - Excused",
    "Absence - Planned Hours - Excused",
    "Absence - PTO Unplanned Hours - Excused",
    "Leave",
    "Actual Hours",
    "Absence  - Unplanned Hours - Unexcused",
]

# Columns overwritten from the new file when they change.
ATTRIBUTE_COLUMNS = [
    "Employment Status",
    "Employee Pay Rule",
    "Temp Agency Code",
    "Temp Hourly Rates",
    "Rehire Date",
    "Profit Center",
    "Shift",
    "Department",
    "Reports To",
    "Job",
]

# Columns forming the match key. Never overwritten.
KEY_COLUMNS = ["Employee Full Name", "Hire Date"]

DATE_COLUMNS = ["Hire Date", "Rehire Date"]

# Helper key in column W: =CONCATENATE(A,B,C,H,I,J,K,U)
HELPER_KEY_SOURCE_COLS = ["A", "B", "C", "H", "I", "J", "K", "U"]
HELPER_KEY_COL_INDEX = len(CANONICAL_COLUMNS) + 1  # column W


# ----------------------------------------------------------------------------
# Loading
# ----------------------------------------------------------------------------

def _normalize_header(name) -> str:
    """Trim and collapse whitespace, preserving the intentional double space."""
    if name is None:
        return ""
    return str(name).strip()


def header_fingerprint(name) -> str:
    """Case- and whitespace-insensitive form of a column name.

    Column V is 'Absence  - Unplanned Hours - Unexcused' with a double space,
    a quirk of the source export. Google Sheets collapses runs of whitespace on
    some paste paths, so an exact-match lookup breaks on a header that is
    visually identical. Matching on the fingerprint tolerates that in either
    direction without letting Excused and Unexcused collide.
    """
    return re.sub(r"\s+", " ", str(name or "").strip()).upper()


def resolve_headers(headers: list) -> list:
    """Map raw sheet/file headers onto canonical names where they match."""
    canonical_by_fp = {header_fingerprint(c): c for c in CANONICAL_COLUMNS}
    alias_by_fp = {header_fingerprint(k): v for k, v in COLUMN_ALIASES.items()}

    resolved = []
    for raw in headers:
        fp = header_fingerprint(raw)
        resolved.append(alias_by_fp.get(fp) or canonical_by_fp.get(fp) or _normalize_header(raw))
    return resolved


def _find_header_row(raw: pd.DataFrame, sentinel: str = "Employment Status") -> int:
    """Locate the header row; weekly exports carry metadata lines above it."""
    for idx in range(min(40, len(raw))):
        row = raw.iloc[idx].astype(str).str.strip()
        if (row == sentinel).any():
            return idx
    raise ValueError(
        f"Could not find a header row containing '{sentinel}' in the first 40 rows."
    )


def _extract_metadata(raw: pd.DataFrame, header_row: int) -> dict:
    """Pull the label/value metadata lines that sit above the header."""
    meta = {}
    for idx in range(header_row):
        label = str(raw.iloc[idx, 0]).strip().rstrip(":")
        if not label or label.lower() == "nan":
            continue
        value = ""
        if raw.shape[1] > 1:
            value = str(raw.iloc[idx, 1]).strip()
            if value.lower() == "nan":
                value = ""
        meta[label] = value
    return meta


def load_table(uploaded_file) -> tuple[pd.DataFrame, dict]:
    """Read an uploaded .csv/.xlsx into a canonical DataFrame plus its metadata."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)

    if name.endswith(".csv"):
        # The export's metadata lines have 2 fields while the header has 22,
        # which defeats pandas' tokenizer. Read rows manually and pad.
        text = uploaded_file.read()
        if isinstance(text, bytes):
            text = text.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            raise ValueError("The file appears to be empty.")
        width = max(len(r) for r in rows)
        rows = [r + [""] * (width - len(r)) for r in rows]
        raw = pd.DataFrame(rows).replace("", pd.NA)
    else:
        raw = pd.read_excel(uploaded_file, header=None, dtype=object)

    header_row = _find_header_row(raw)
    metadata = _extract_metadata(raw, header_row)

    headers = resolve_headers(raw.iloc[header_row].tolist())
    df = raw.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)

    # Drop unnamed/helper trailing columns (the CONCATENATE and COUNTIF columns).
    df = df.loc[:, [c for c in df.columns if c != ""]]
    df = df.loc[:, ~df.columns.duplicated()]

    missing = [c for c in CANONICAL_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"'{uploaded_file.name}' is missing expected column(s): {', '.join(missing)}"
        )

    df = df[CANONICAL_COLUMNS].copy()

    # Drop rows with no name and no hire date - trailing blanks from the export.
    blank = df["Employee Full Name"].isna() & df["Hire Date"].isna()
    df = df[~blank].reset_index(drop=True)

    for col in HOURS_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in DATE_COLUMNS:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    return df, metadata


# ----------------------------------------------------------------------------
# Keys
# ----------------------------------------------------------------------------

def build_key(df: pd.DataFrame) -> pd.Series:
    """Name + Hire Date, case- and whitespace-insensitive."""
    name = (
        df["Employee Full Name"].fillna("").astype(str)
        .str.strip().str.upper().str.replace(r"\s+", " ", regex=True)
    )
    hire = df["Hire Date"].dt.strftime("%Y-%m-%d").fillna("")
    return name + " | " + hire


def collapse_duplicates(df: pd.DataFrame, keys: pd.Series) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Sum hours across duplicate keys; keep the last non-null attribute values."""
    df = df.copy()
    df["__key"] = keys
    dupe_mask = df["__key"].duplicated(keep=False)
    dupes = df[dupe_mask].drop(columns="__key").copy()
    if not dupe_mask.any():
        return df.drop(columns="__key"), dupes

    agg = {}
    for col in df.columns:
        if col == "__key":
            continue
        agg[col] = "sum" if col in HOURS_COLUMNS else "last"

    collapsed = (
        df.groupby("__key", sort=False, dropna=False)
        .agg(agg)
        .reset_index(drop=True)
    )
    # groupby.sum turns all-NaN groups into 0; restore genuine blanks.
    for col in HOURS_COLUMNS:
        all_nan = df.groupby("__key", sort=False, dropna=False)[col].apply(
            lambda s: s.isna().all()
        ).reset_index(drop=True)
        collapsed.loc[all_nan.values, col] = pd.NA
        collapsed[col] = pd.to_numeric(collapsed[col], errors="coerce")

    return collapsed, dupes


# ----------------------------------------------------------------------------
# Comparison helpers
# ----------------------------------------------------------------------------

def values_differ(old, new) -> bool:
    """True when a field genuinely changed, ignoring blank/format noise."""
    old_blank = pd.isna(old) or str(old).strip() == ""
    new_blank = pd.isna(new) or str(new).strip() == ""
    if old_blank and new_blank:
        return False
    if old_blank != new_blank:
        return True

    if isinstance(old, (pd.Timestamp, datetime)) or isinstance(new, (pd.Timestamp, datetime)):
        o = pd.to_datetime(old, errors="coerce")
        n = pd.to_datetime(new, errors="coerce")
        if pd.notna(o) and pd.notna(n):
            return o.normalize() != n.normalize()

    try:
        return abs(float(old) - float(new)) > 1e-9
    except (TypeError, ValueError):
        pass

    return str(old).strip() != str(new).strip()


def display_value(v) -> str:
    if pd.isna(v) or str(v).strip() == "":
        return "(blank)"
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%-m/%-d/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


# ----------------------------------------------------------------------------
# Merge
# ----------------------------------------------------------------------------

def normalized_name(value) -> str:
    """Employee name stripped of case and whitespace noise, for rehire linking."""
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def link_rehires(hist, new, dropped_keys, new_hire_keys):
    """Reconnect employees whose Hire Date was rewritten between exports.

    A person absent from this week's export whose name reappears under a new
    hire date is a rehire, not a departure plus a new hire. Their cumulative
    hours must carry forward.

    Pairs 1:1 only. When a name has several candidates on either side the
    pairing is genuinely ambiguous, so it is flagged for manual review rather
    than guessed at - misassigning hours between two people who share a name
    is worse than leaving it for a human.
    """
    dropped_by_name, new_by_name = {}, {}

    for key in dropped_keys:
        name = normalized_name(hist.loc[key, "Employee Full Name"])
        if name and name != "TEMP-":
            dropped_by_name.setdefault(name, []).append(key)

    for key in new_hire_keys:
        name = normalized_name(new.loc[key, "Employee Full Name"])
        if name and name != "TEMP-":
            new_by_name.setdefault(name, []).append(key)

    pairs, ambiguous = [], []
    for name in set(dropped_by_name) & set(new_by_name):
        old_keys, fresh_keys = dropped_by_name[name], new_by_name[name]
        if len(old_keys) == 1 and len(fresh_keys) == 1:
            pairs.append((old_keys[0], fresh_keys[0]))
        else:
            ambiguous.append({
                "Employee Full Name": new.loc[fresh_keys[0], "Employee Full Name"],
                "Rows in historical": len(old_keys),
                "Rows in export": len(fresh_keys),
                "Cumulative hours at risk": round(
                    float(hist.loc[old_keys, "Actual Hours"].sum(skipna=True)), 2
                ),
            })
    return pairs, ambiguous


def sync_duplicate_rows(df: pd.DataFrame, fields: list, temp_only: bool = False):
    """Make every row for one person agree on person-level fields.

    Someone with several stints has several rows. Facts about the person -
    chiefly Employment Status - should not disagree between them, so the most
    recent row wins and the older rows are brought into line. No row is removed
    and no hours move; each row keeps the hours it earned.
    """
    if not fields:
        return df, pd.DataFrame(), pd.DataFrame()

    df = df.copy().reset_index(drop=True)
    names = df["Employee Full Name"].apply(normalized_name)

    # Rank rows within a person: newest hire date wins, then rehire date,
    # then file order. Missing dates sort oldest.
    order = pd.DataFrame({
        "hire": df["Hire Date"],
        "rehire": df["Rehire Date"],
        "pos": range(len(df)),
    })

    sync_log, groups_seen = [], []

    for name, idx in names.groupby(names).groups.items():
        if not name or name == "TEMP-" or len(idx) < 2:
            continue
        if temp_only and not name.startswith("TEMP-"):
            continue

        block = order.loc[idx].sort_values(
            ["hire", "rehire", "pos"], na_position="first"
        )
        latest_i = block.index[-1]
        older = list(block.index[:-1])

        changed_here = 0
        for col in fields:
            authority = df.at[latest_i, col]
            for i in older:
                if values_differ(df.at[i, col], authority):
                    sync_log.append({
                        "Employee Full Name": df.at[i, "Employee Full Name"],
                        "Row hire date": display_value(df.at[i, "Hire Date"]),
                        "Field": col,
                        "Was": display_value(df.at[i, col]),
                        "Now": display_value(authority),
                        "Matched to hire date": display_value(df.at[latest_i, "Hire Date"]),
                        "TEMP": "Yes" if name.startswith("TEMP-") else "No",
                    })
                    df.at[i, col] = authority
                    changed_here += 1

        groups_seen.append({
            "Employee Full Name": df.at[latest_i, "Employee Full Name"],
            "Rows": len(idx),
            "Latest hire date": display_value(df.at[latest_i, "Hire Date"]),
            "Fields changed": changed_here,
            "TEMP": "Yes" if name.startswith("TEMP-") else "No",
        })

    return df, pd.DataFrame(sync_log), pd.DataFrame(groups_seen)


def newest_key_per_person(hist, new, temp_only: bool = False) -> set:
    """Keys that are their person's most recent row.

    Used to keep the weekly attribute overwrite off older rows for any field
    the duplicate sync owns. Without this the two stages fight every week: the
    export reports an old stint as Terminated, the sync reports the person as
    Active, and the change log fills with a flip that nets to nothing.
    """
    frames = [
        hist[["Employee Full Name", "Hire Date", "Rehire Date"]],
        new.loc[[k for k in new.index if k not in hist.index],
                ["Employee Full Name", "Hire Date", "Rehire Date"]],
    ]
    cand = pd.concat(frames)
    cand = cand[~cand.index.duplicated(keep="first")]
    cand = cand.assign(
        __name=cand["Employee Full Name"].apply(normalized_name),
        __pos=range(len(cand)),
    )

    synced = cand["__name"].ne("") & cand["__name"].ne("TEMP-")
    if temp_only:
        synced &= cand["__name"].str.startswith("TEMP-")

    # Anything the sync won't touch keeps its normal update path.
    newest = set(cand.index[~synced])

    considered = cand[synced].sort_values(
        ["__name", "Hire Date", "Rehire Date", "__pos"], na_position="first"
    )
    newest |= set(considered.groupby("__name", sort=False).tail(1).index)
    return newest


def merge_rosters(hist: pd.DataFrame, new: pd.DataFrame, drop_missing: bool,
                  link_rehire: bool = True, sync_fields: list = None,
                  sync_temp_only: bool = False) -> dict:
    hist_keys = build_key(hist)
    new_keys = build_key(new)

    hist, hist_dupes = collapse_duplicates(hist, hist_keys)
    new, new_dupes = collapse_duplicates(new, new_keys)

    hist_keys = build_key(hist)
    new_keys = build_key(new)

    hist = hist.set_index(hist_keys)
    new = new.set_index(new_keys)

    matched = [k for k in hist.index if k in new.index]
    new_hire_keys = [k for k in new.index if k not in hist.index]
    dropped_keys = [k for k in hist.index if k not in new.index]

    # Fields the duplicate sync owns must not be written onto older rows.
    sync_owned = set(sync_fields or [])
    newest_keys = (
        newest_key_per_person(hist, new, temp_only=sync_temp_only)
        if sync_owned else set(hist.index) | set(new.index)
    )

    result = hist.copy()
    hours_log, field_log, deferred_log = [], [], []

    if matched:
        # --- Hours, vectorized ---------------------------------------------
        # Elementwise add over aligned indexes. A per-row Python loop here cost
        # seconds on every rerun, which reads as the app freezing.
        h_hours = hist.loc[matched, HOURS_COLUMNS]
        n_hours = new.loc[matched, HOURS_COLUMNS].apply(pd.to_numeric, errors="coerce")
        h_hours = h_hours.apply(pd.to_numeric, errors="coerce")

        both_blank = h_hours.isna() & n_hours.isna()
        totals = (h_hours.fillna(0) + n_hours.fillna(0)).mask(both_blank)
        result.loc[matched, HOURS_COLUMNS] = totals

        contributed = n_hours.fillna(0).ne(0)
        rows_with_hours = contributed.any(axis=1)
        for key in n_hours.index[rows_with_hours]:
            added = {c: round(float(n_hours.at[key, c]), 2)
                     for c in HOURS_COLUMNS if contributed.at[key, c]}
            hours_log.append({
                "Employee Full Name": hist.at[key, "Employee Full Name"],
                "Hire Date": display_value(hist.at[key, "Hire Date"]),
                **added,
                "Hours Added (Actual)": round(
                    float(n_hours.at[key, "Actual Hours"] or 0), 2
                ),
            })

        # --- Attributes ------------------------------------------------------
        # values_differ is careful but slow. A plain string comparison flags a
        # superset of real changes, so it cheaply narrows the candidates and
        # values_differ only runs on those.
        h_attr = hist.loc[matched, ATTRIBUTE_COLUMNS]
        n_attr = new.loc[matched, ATTRIBUTE_COLUMNS]
        candidates = h_attr.astype(str).values != n_attr.astype(str).values

        for r_i, key in enumerate(matched):
            for c_i, col in enumerate(ATTRIBUTE_COLUMNS):
                if not candidates[r_i, c_i]:
                    continue
                old_v, new_v = h_attr.iat[r_i, c_i], n_attr.iat[r_i, c_i]
                if not values_differ(old_v, new_v):
                    continue

                if col in sync_owned and key not in newest_keys:
                    # An older stint. The sync sets this field from the person's
                    # newest row, so writing the export's value here would only
                    # be undone a moment later.
                    deferred_log.append({
                        "Employee Full Name": hist.at[key, "Employee Full Name"],
                        "Hire Date": display_value(hist.at[key, "Hire Date"]),
                        "Field": col,
                        "Export says": display_value(new_v),
                        "Left as": display_value(old_v),
                    })
                    continue

                result.at[key, col] = new_v
                field_log.append({
                    "Employee Full Name": hist.at[key, "Employee Full Name"],
                    "Hire Date": display_value(hist.at[key, "Hire Date"]),
                    "Field": col,
                    "Was": display_value(old_v),
                    "Now": display_value(new_v),
                })

    new_hires = new.loc[new_hire_keys].copy() if new_hire_keys else new.iloc[0:0].copy()
    dropped = hist.loc[dropped_keys].copy() if dropped_keys else hist.iloc[0:0].copy()

    # --- Rehire reconciliation ---------------------------------------------
    # Someone whose Hire Date was rewritten looks like a departure plus a new
    # hire. Relink them so their cumulative hours survive.
    rehire_log, ambiguous = [], []
    rehire_rows = {}

    if link_rehire:
        pairs, ambiguous = link_rehires(hist, new, dropped_keys, new_hire_keys)

        for old_key, fresh_key in pairs:
            h_row, n_row = hist.loc[old_key], new.loc[fresh_key]
            merged_row = n_row.copy()  # adopt current attributes and hire date

            for col in HOURS_COLUMNS:
                old_v, new_v = h_row[col], n_row[col]
                if pd.isna(old_v) and pd.isna(new_v):
                    merged_row[col] = pd.NA
                else:
                    merged_row[col] = (0 if pd.isna(old_v) else float(old_v)) + \
                                      (0 if pd.isna(new_v) else float(new_v))

            rehire_rows[fresh_key] = merged_row
            rehire_log.append({
                "Employee Full Name": n_row["Employee Full Name"],
                "Hire Date was": display_value(h_row["Hire Date"]),
                "Hire Date now": display_value(n_row["Hire Date"]),
                "Status was": display_value(h_row["Employment Status"]),
                "Status now": display_value(n_row["Employment Status"]),
                "Hours carried forward": round(
                    0 if pd.isna(h_row["Actual Hours"]) else float(h_row["Actual Hours"]), 2
                ),
                "Hours added this week": round(
                    0 if pd.isna(n_row["Actual Hours"]) else float(n_row["Actual Hours"]), 2
                ),
            })

        paired_old = {o for o, _ in pairs}
        paired_new = {f for _, f in pairs}

        # A relinked person is no longer a departure or a new hire.
        dropped_keys = [k for k in dropped_keys if k not in paired_old]
        new_hire_keys = [k for k in new_hire_keys if k not in paired_new]
        new_hires = new.loc[new_hire_keys].copy() if new_hire_keys else new.iloc[0:0].copy()
        dropped = hist.loc[dropped_keys].copy() if dropped_keys else hist.iloc[0:0].copy()

        # The superseded historical row always goes, toggle or not.
        if paired_old:
            result = result.drop(index=list(paired_old))

    if drop_missing and dropped_keys:
        result = result.drop(index=dropped_keys)

    if rehire_rows:
        result = pd.concat([result, pd.DataFrame(rehire_rows).T])

    if new_hire_keys:
        result = pd.concat([result, new_hires])

    result = result.reset_index(drop=True)
    for col in HOURS_COLUMNS:
        result[col] = pd.to_numeric(result[col], errors="coerce")
    for col in DATE_COLUMNS:
        result[col] = pd.to_datetime(result[col], errors="coerce")

    rows_before_sync = len(result)
    result, sync_log, sync_groups = sync_duplicate_rows(
        result, sync_fields or [], temp_only=sync_temp_only
    )
    assert len(result) == rows_before_sync, "sync must never add or remove rows"

    return {
        "result": result,
        "hours_log": pd.DataFrame(hours_log),
        "field_log": pd.DataFrame(field_log),
        "rehire_log": pd.DataFrame(rehire_log),
        "sync_log": sync_log,
        "deferred_log": pd.DataFrame(deferred_log),
        "sync_groups": sync_groups,
        "ambiguous": pd.DataFrame(ambiguous),
        "new_hires": new_hires.reset_index(drop=True),
        "dropped": dropped.reset_index(drop=True),
        "hist_dupes": hist_dupes,
        "new_dupes": new_dupes,
        "matched_count": len(matched),
    }


# ----------------------------------------------------------------------------
# Excel output
# ----------------------------------------------------------------------------

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Employee Hours Test") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    for col_idx, header in enumerate(CANONICAL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(CANONICAL_COLUMNS, start=1):
            value = row[col]
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font
            if col in DATE_COLUMNS:
                cell.number_format = "m/d/yyyy"
            elif col in HOURS_COLUMNS:
                cell.number_format = "#,##0.00"

        refs = ",".join(f"{letter}{r}" for letter in HELPER_KEY_SOURCE_COLS)
        helper = ws.cell(row=r, column=HELPER_KEY_COL_INDEX, value=f"=CONCATENATE({refs})")
        helper.font = body_font

    widths = {"A": 18, "B": 28, "C": 14, "D": 16, "E": 14, "F": 12, "G": 12,
              "H": 13, "I": 20, "J": 12, "K": 22, "U": 28}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for i in range(12, 23):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions[get_column_letter(HELPER_KEY_COL_INDEX)].width = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CANONICAL_COLUMNS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------------
# Google Sheets persistence
# ----------------------------------------------------------------------------
# Spreadsheet tabs:
#   Historical   Current cumulative roster used by the app/download button
#   Baseline     One-time snapshot of Historical when this version went live
#   Weekly_Data  Raw canonical rows from every post-go-live weekly upload
#   Update_Log   One row per processed week, including hash/status/audit fields
#
# Baseline + Weekly_Data make replacement safe: if an old week is corrected,
# the app replays every stored week in sequence through THE SAME merge_rosters()
# function instead of trying to subtract data from cumulative totals.
#
# A brand-new week does NOT replay. Historical already equals the replay of
# every prior week, so merging the new week straight into it is identical and
# runs in constant time however many weeks are stored.

HISTORICAL_SHEET = "Historical"
BASELINE_SHEET = "Baseline"
WEEKLY_DATA_SHEET = "Weekly_Data"
UPDATE_LOG_SHEET = "Update_Log"

WEEKLY_META_COLUMNS = [
    "__Period Key",
    "__Period",
    "__Sequence",
    "__File Hash",
    "__Source Filename",
]
WEEKLY_STORAGE_COLUMNS = WEEKLY_META_COLUMNS + CANONICAL_COLUMNS

LOG_COLUMNS = [
    "Period Key",
    "Period",
    "Sequence",
    "File Hash",
    "Source Filename",
    "Status",
    "Processed At UTC",
    "Rows",
    "New Hires",
    "Active to Terminated",
    "Terminated to Active",
    "Actual Hours Added",
]


def file_sha256(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def normalize_period_display(period) -> str:
    return re.sub(r"\s+", " ", str(period or "").strip())


def canonical_period_key(period) -> str:
    """Stable key for duplicate detection; prefers the two dates in Time Period."""
    display = normalize_period_display(period)
    date_tokens = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", display)
    parsed = []
    for token in date_tokens[:2]:
        dt = pd.to_datetime(token, errors="coerce")
        if pd.notna(dt):
            parsed.append(dt.normalize())
    if len(parsed) >= 2:
        return f"{parsed[0]:%Y-%m-%d}__{parsed[1]:%Y-%m-%d}"
    return re.sub(r"\s+", " ", display.upper())


def _sheet_cell_value(value, column=None):
    """Convert pandas/numpy values into JSON-safe Google Sheets cell values."""
    if pd.isna(value):
        return ""
    if column in DATE_COLUMNS:
        dt = pd.to_datetime(value, errors="coerce")
        return "" if pd.isna(dt) else dt.strftime("%Y-%m-%d")
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (ValueError, AttributeError):
            pass
    return value


def dataframe_to_sheet_rows(df: pd.DataFrame, columns: list) -> list[list]:
    rows = []
    for _, row in df.iterrows():
        rows.append([_sheet_cell_value(row.get(col, ""), col) for col in columns])
    return rows


def roster_from_sheet_values(values: list[list], source_name: str) -> pd.DataFrame:
    """Read a Google Sheet table into the same canonical dataframe as load_table."""
    if not values:
        return pd.DataFrame(columns=CANONICAL_COLUMNS)

    headers = resolve_headers(values[0])

    missing = [c for c in CANONICAL_COLUMNS if c not in headers]
    if missing:
        raise ValueError(
            f"Google Sheet tab '{source_name}' is missing expected column(s): "
            + ", ".join(missing)
        )

    rows = values[1:]
    width = len(headers)
    padded = [r + [""] * (width - len(r)) for r in rows]
    df = pd.DataFrame(padded, columns=headers)
    df = df.loc[:, ~df.columns.duplicated()]
    df = df[CANONICAL_COLUMNS].copy()
    df = df.replace("", pd.NA)

    blank = df["Employee Full Name"].isna() & df["Hire Date"].isna()
    df = df[~blank].reset_index(drop=True)

    for col in HOURS_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in DATE_COLUMNS:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    return df


@st.cache_resource(show_spinner=False)
def get_google_spreadsheet():
    """Authenticate with a service account stored in Streamlit secrets."""
    try:
        credentials = dict(st.secrets["gcp_service_account"])
        sheet_id = str(st.secrets["GOOGLE_SHEET_ID"]).strip()
    except Exception as exc:
        raise RuntimeError(
            "Google Sheets secrets are not configured. Add GOOGLE_SHEET_ID and "
            "[gcp_service_account] to Streamlit secrets."
        ) from exc

    # TOML often stores literal \\n sequences; Google expects real newlines.
    if "private_key" in credentials:
        credentials["private_key"] = credentials["private_key"].replace("\\n", "\n")

    client = gspread.service_account_from_dict(credentials)
    return client.open_by_key(sheet_id)


# --- PATCH 1: cached tab reads -----------------------------------------------
# Every rerun previously called get_all_values() on four tabs over the network.
# Values are now cached and invalidated explicitly after a successful commit.

def sheet_version() -> int:
    """Bumped after each commit so cached reads are invalidated on demand."""
    return st.session_state.get("_sheet_version", 0)


def bump_sheet_cache():
    st.session_state["_sheet_version"] = sheet_version() + 1


@st.cache_data(ttl=900, show_spinner=False)
def fetch_tab_values(title: str, version: int) -> list[list]:
    """Read one tab. `version` is part of the cache key, not used in the body."""
    book = get_google_spreadsheet()
    try:
        return book.worksheet(title).get_all_values()
    except gspread.WorksheetNotFound:
        return []


def get_or_create_worksheet(book, title: str, rows: int, cols: int):
    try:
        return book.worksheet(title)
    except gspread.WorksheetNotFound:
        return book.add_worksheet(title=title, rows=rows, cols=cols)


def ensure_sheet_header(ws, header: list[str]):
    first_row = ws.row_values(1)
    if not first_row:
        if ws.col_count < len(header):
            ws.resize(cols=len(header))
        ws.update([header], "A1", raw=True)
        return

    normalized = [header_fingerprint(v) for v in first_row[:len(header)]]
    if normalized != [header_fingerprint(h) for h in header]:
        raise ValueError(
            f"Google Sheet tab '{ws.title}' has an unexpected header. "
            "Do not rename/reorder its system columns."
        )


def read_roster_worksheet(ws) -> pd.DataFrame:
    return roster_from_sheet_values(ws.get_all_values(), ws.title)


def write_roster_worksheet(ws, df: pd.DataFrame):
    values = [CANONICAL_COLUMNS] + dataframe_to_sheet_rows(df, CANONICAL_COLUMNS)
    needed_rows = max(1000, len(values) + 20)
    needed_cols = max(26, len(CANONICAL_COLUMNS))
    if ws.row_count < needed_rows or ws.col_count < needed_cols:
        ws.resize(rows=max(ws.row_count, needed_rows), cols=max(ws.col_count, needed_cols))
    ws.clear()
    ws.update(values, "A1", raw=True)
    ws.freeze(rows=1)


def read_update_log(ws) -> pd.DataFrame:
    values = ws.get_all_values()
    if len(values) <= 1:
        return pd.DataFrame(columns=LOG_COLUMNS)
    width = len(LOG_COLUMNS)
    rows = [r + [""] * (width - len(r)) for r in values[1:]]
    return pd.DataFrame([r[:width] for r in rows], columns=LOG_COLUMNS)


def find_log_row(log_df: pd.DataFrame, period_key: str):
    if log_df.empty:
        return None
    matches = log_df.index[log_df["Period Key"].astype(str) == period_key].tolist()
    return None if not matches else log_df.loc[matches[-1]].to_dict()


def next_sequence(log_df: pd.DataFrame) -> int:
    if log_df.empty:
        return 1
    seq = pd.to_numeric(log_df["Sequence"], errors="coerce")
    return 1 if seq.dropna().empty else int(seq.max()) + 1


def upsert_log_row(ws, record: dict):
    values = ws.get_all_values()
    row_number = None
    for i, row in enumerate(values[1:], start=2):
        if row and str(row[0]) == str(record["Period Key"]):
            row_number = i
            break

    row_values = [_sheet_cell_value(record.get(col, "")) for col in LOG_COLUMNS]
    last_col = get_column_letter(len(LOG_COLUMNS))
    if row_number is None:
        ws.append_row(row_values, value_input_option="RAW")
    else:
        ws.update([row_values], f"A{row_number}:{last_col}{row_number}", raw=True)


def _contiguous_blocks(row_numbers: list[int]) -> list[tuple[int, int]]:
    if not row_numbers:
        return []
    nums = sorted(row_numbers)
    blocks = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        blocks.append((start, prev))
        start = prev = n
    blocks.append((start, prev))
    return blocks


def archive_weekly_period(ws, new_df: pd.DataFrame, period_key: str, period: str,
                          sequence: int, file_hash: str, filename: str):
    """Replace this period's archived rows in-place, or append it if it is new."""
    existing = ws.get_all_values()
    rows_for_period = [
        i for i, row in enumerate(existing[1:], start=2)
        if row and str(row[0]) == period_key
    ]
    insert_at = min(rows_for_period) if rows_for_period else None

    # Delete bottom-up so row numbers above each deleted block remain valid.
    for start, end in reversed(_contiguous_blocks(rows_for_period)):
        ws.delete_rows(start, end)

    archive_rows = []
    for _, row in new_df.iterrows():
        meta = [period_key, period, sequence, file_hash, filename]
        canonical = [_sheet_cell_value(row[col], col) for col in CANONICAL_COLUMNS]
        archive_rows.append(meta + canonical)

    if not archive_rows:
        raise ValueError("The weekly export contains no employee rows to archive.")

    if insert_at is None:
        ws.append_rows(archive_rows, value_input_option="RAW")
    else:
        ws.insert_rows(
            archive_rows,
            row=insert_at,
            value_input_option="RAW",
            inherit_from_before=(insert_at > 1),
        )


def weekly_groups_from_archive(ws) -> list[dict]:
    values = ws.get_all_values()
    if len(values) <= 1:
        return []

    width = len(WEEKLY_STORAGE_COLUMNS)
    rows = [r + [""] * (width - len(r)) for r in values[1:]]
    archive = pd.DataFrame([r[:width] for r in rows], columns=WEEKLY_STORAGE_COLUMNS)
    archive["__Sequence"] = pd.to_numeric(archive["__Sequence"], errors="coerce")
    archive = archive[archive["__Period Key"].astype(str).str.strip() != ""].copy()

    groups = []
    for (sequence, period_key), block in archive.groupby(
        ["__Sequence", "__Period Key"], sort=True, dropna=False
    ):
        if pd.isna(sequence):
            continue

        week = block[CANONICAL_COLUMNS].copy().replace("", pd.NA)
        for col in HOURS_COLUMNS:
            week[col] = pd.to_numeric(week[col], errors="coerce")
        for col in DATE_COLUMNS:
            week[col] = pd.to_datetime(week[col], errors="coerce")

        groups.append({
            "sequence": int(sequence),
            "period_key": str(period_key),
            "period": str(block["__Period"].iloc[0]),
            "file_hash": str(block["__File Hash"].iloc[0]),
            "filename": str(block["__Source Filename"].iloc[0]),
            "df": week.reset_index(drop=True),
        })

    groups.sort(key=lambda x: x["sequence"])
    return groups


def sort_historical_output(df: pd.DataFrame) -> pd.DataFrame:
    if not SORT_OUTPUT:
        return df.reset_index(drop=True)
    return df.sort_values(
        "Employee Full Name", key=lambda s: s.astype(str).str.upper()
    ).reset_index(drop=True)


# --- PATCH 3a: replay now uses the cached merge -------------------------------

def rebuild_from_baseline(baseline_df: pd.DataFrame, weekly_ws, target_period_key=None):
    """Replay every archived week. Only needed when correcting a past week."""
    current = baseline_df.copy()
    target_merge = None
    replay_ambiguous = []

    for group in weekly_groups_from_archive(weekly_ws):
        merged = merge_cached(
            current,
            group["df"],
            DROP_MISSING,
            LINK_REHIRE,
            tuple(SYNC_FIELDS),
            SYNC_TEMP_ONLY,
        )
        current = sort_historical_output(merged["result"])

        if len(merged["ambiguous"]):
            tmp = merged["ambiguous"].copy()
            tmp.insert(0, "Period", group["period"])
            replay_ambiguous.append(tmp)

        if group["period_key"] == target_period_key:
            target_merge = merged

    all_ambiguous = (
        pd.concat(replay_ambiguous, ignore_index=True)
        if replay_ambiguous else pd.DataFrame()
    )
    return current, target_merge, all_ambiguous


def count_people(log, was, now):
    """Distinct people whose Employment Status moved from `was` to `now`."""
    names = set()
    for frame in log:
        if not len(frame):
            continue
        if not {"Field", "Was", "Now"} <= set(frame.columns):
            continue
        hit = frame[
            (frame["Field"] == "Employment Status")
            & (frame["Was"] == was)
            & (frame["Now"] == now)
        ]
        names |= {normalized_name(v) for v in hit["Employee Full Name"]}
    names.discard("")
    return len(names)


def merge_summary(merged: dict, weekly_df: pd.DataFrame) -> dict:
    terminated = count_people(
        [merged["field_log"], merged["sync_log"]], "Active", "Terminated"
    )
    reactivated = count_people(
        [merged["field_log"], merged["sync_log"]], "Terminated", "Active"
    )
    actual_added = pd.to_numeric(weekly_df["Actual Hours"], errors="coerce").sum(skipna=True)
    return {
        "Rows": len(weekly_df),
        "New Hires": len(merged["new_hires"]),
        "Active to Terminated": terminated,
        "Terminated to Active": reactivated,
        "Actual Hours Added": round(float(actual_added), 2),
    }


# --- PATCH 2: cached reads, Weekly_Data no longer read at startup -------------

def initialize_google_backend():
    """Create system tabs and snapshot Historical -> Baseline exactly once."""
    book = get_google_spreadsheet()
    historical_ws = get_or_create_worksheet(book, HISTORICAL_SHEET, rows=5000, cols=26)
    baseline_ws = get_or_create_worksheet(book, BASELINE_SHEET, rows=5000, cols=26)
    weekly_ws = get_or_create_worksheet(
        book, WEEKLY_DATA_SHEET, rows=5000, cols=len(WEEKLY_STORAGE_COLUMNS)
    )
    log_ws = get_or_create_worksheet(book, UPDATE_LOG_SHEET, rows=500, cols=len(LOG_COLUMNS))

    ensure_sheet_header(weekly_ws, WEEKLY_STORAGE_COLUMNS)
    ensure_sheet_header(log_ws, LOG_COLUMNS)

    v = sheet_version()
    historical_values = fetch_tab_values(HISTORICAL_SHEET, v)
    baseline_values = fetch_tab_values(BASELINE_SHEET, v)

    # Historical/Baseline are allowed to be brand-new blank tabs.
    historical_df = (
        roster_from_sheet_values(historical_values, HISTORICAL_SHEET)
        if historical_values else pd.DataFrame(columns=CANONICAL_COLUMNS)
    )
    baseline_df = (
        roster_from_sheet_values(baseline_values, BASELINE_SHEET)
        if baseline_values else pd.DataFrame(columns=CANONICAL_COLUMNS)
    )

    if baseline_df.empty:
        if historical_df.empty:
            raise RuntimeError(
                "The Google Sheet is connected, but the 'Historical' tab is empty. "
                "For the one-time migration, paste/import your CURRENT historical "
                "roster into the Historical tab with the same A-V headers, then rerun."
            )
        write_roster_worksheet(baseline_ws, historical_df)
        bump_sheet_cache()
        baseline_df = historical_df.copy()

    # If Historical was accidentally cleared after Baseline existed, restore it
    # deterministically from Baseline + archived weeks.
    if historical_df.empty:
        rebuilt, _, _ = rebuild_from_baseline(baseline_df, weekly_ws)
        write_roster_worksheet(historical_ws, rebuilt)
        bump_sheet_cache()
        historical_df = rebuilt

    return {
        "book": book,
        "historical_ws": historical_ws,
        "baseline_ws": baseline_ws,
        "weekly_ws": weekly_ws,
        "log_ws": log_ws,
        "historical_df": historical_df,
        "baseline_df": baseline_df,
    }


# --- PATCH 3b: append for a new week, replay only for corrections -------------

def process_week_into_google(backend: dict, new_df: pd.DataFrame, period: str,
                             period_key: str, file_hash: str, filename: str,
                             sequence: int, mode: str = "append"):
    """Archive one week, update Historical, commit the audit log last.

    mode="append"  a brand-new latest week. Historical already equals the replay
                   of every prior week, so merging this week straight into it
                   gives the identical result in constant time however many
                   weeks are stored.
    mode="replay"  correcting or resuming a week. Rebuild from Baseline through
                   the whole archive, because earlier weeks may now differ.
    """
    log_ws = backend["log_ws"]
    weekly_ws = backend["weekly_ws"]
    historical_ws = backend["historical_ws"]

    pending = {
        "Period Key": period_key,
        "Period": period,
        "Sequence": sequence,
        "File Hash": file_hash,
        "Source Filename": filename,
        "Status": "PENDING",
        "Processed At UTC": "",
        "Rows": len(new_df),
        "New Hires": "",
        "Active to Terminated": "",
        "Terminated to Active": "",
        "Actual Hours Added": "",
    }
    upsert_log_row(log_ws, pending)

    # Always replace the period's archive rows, whichever mode. That makes a
    # PENDING transaction safely resumable and prevents duplicate raw rows.
    archive_weekly_period(
        weekly_ws, new_df, period_key, period, sequence, file_hash, filename
    )

    if mode == "append":
        target_merge = merge_cached(
            backend["historical_df"], new_df,
            DROP_MISSING, LINK_REHIRE, tuple(SYNC_FIELDS), SYNC_TEMP_ONLY,
        )
        final_df = sort_historical_output(target_merge["result"])
        replay_ambiguous = pd.DataFrame()
    else:
        final_df, target_merge, replay_ambiguous = rebuild_from_baseline(
            backend["baseline_df"], weekly_ws, target_period_key=period_key
        )
        if target_merge is None:
            raise RuntimeError("The uploaded period could not be found after archiving.")

    write_roster_worksheet(historical_ws, final_df)

    summary = merge_summary(target_merge, new_df)
    committed = {
        **pending,
        "Status": "COMMITTED",
        "Processed At UTC": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        **summary,
    }
    upsert_log_row(log_ws, committed)

    bump_sheet_cache()
    backend["historical_df"] = final_df
    return final_df, target_merge, replay_ambiguous, committed


# ----------------------------------------------------------------------------
# Cached wrappers
# ----------------------------------------------------------------------------

@st.cache_data(show_spinner=False, max_entries=4)
def load_table_cached(file_bytes: bytes, filename: str):
    buf = io.BytesIO(file_bytes)
    buf.name = filename
    return load_table(buf)


@st.cache_data(show_spinner="Merging rosters...", max_entries=64)
def merge_cached(hist_df, new_df, drop_missing, link_rehire, sync_fields, sync_temp_only):
    return merge_rosters(
        hist_df, new_df,
        drop_missing=drop_missing,
        link_rehire=link_rehire,
        sync_fields=list(sync_fields),
        sync_temp_only=sync_temp_only,
    )


@st.cache_data(show_spinner="Building workbook...", max_entries=4)
def build_excel_cached(df):
    return to_excel_bytes(df)


# ----------------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------------

st.set_page_config(page_title="Employee Hours Roster Updater", page_icon="", layout="wide")

st.title("Employee Hours Roster Updater")
st.caption(
    "Upload only this week's Employee Hours export. The cumulative historical "
    "roster is stored and updated automatically in Google Sheets. Employees are "
    "matched on **Employee Full Name + Hire Date**."
)

# Fixed behaviour. These were toggles; they are now always on.
LINK_REHIRE = True
DROP_MISSING = True
SORT_OUTPUT = True
SYNC_FIELDS = list(ATTRIBUTE_COLUMNS)   # every attribute field
SYNC_TEMP_ONLY = False                  # applies to everyone, not just TEMPs

link_rehire = LINK_REHIRE
drop_missing = DROP_MISSING
sort_output = SORT_OUTPUT
sync_fields = SYNC_FIELDS
sync_temp_only = SYNC_TEMP_ONLY

with st.sidebar:
    st.header("How this runs")

    st.markdown(
        "**Historical storage**  \n"
        "Google Sheets is the source of truth. You no longer upload last week's "
        "historical file."
    )
    st.markdown(
        "**Matching**  \n"
        "Employee Full Name + Hire Date."
    )
    st.markdown(
        "**Hours**  \n"
        "Every hours column accumulates week over week."
    )
    st.markdown(
        "**Rehires**  \n"
        "A rewritten hire date is treated as a rehire, not a departure plus a "
        "new hire, so cumulative hours carry forward."
    )
    st.markdown(
        "**Missing from the export**  \n"
        "Removed from the file, along with their cumulative hours."
    )
    st.markdown(
        "**Duplicate rows**  \n"
        "When one person has several rows, the newest row is the authority for "
        "every attribute field and older rows are brought into line. No row is "
        "removed and no hours move."
    )

    with st.expander("Fields synced from the newest row"):
        st.write("\n".join(f"- {c}" for c in SYNC_FIELDS))
        st.caption(
            "Because the newest row owns these, the weekly export does not "
            "update them on older rows."
        )

    st.markdown(
        "**Never changed**  \n"
        "Employee Full Name and Hire Date. Output is sorted by name."
    )
    st.markdown(
        "**Duplicate-week protection**  \n"
        "The same exact file is ignored. A different file for an already stored "
        "period requires your confirmation before replacement."
    )

# --- Connect / initialize Google Sheets --------------------------------------
try:
    backend = initialize_google_backend()
except Exception as exc:
    st.error(f"Could not initialize Google Sheets: {exc}")
    st.stop()

historical_df = backend["historical_df"]

st.success(
    f"Google historical database connected — **{len(historical_df):,} rows** currently stored."
)

# The historical download is always available, even before a new weekly upload.
st.subheader("Historical file")
st.download_button(
    "Download current historical file",
    data=build_excel_cached(historical_df),
    file_name="Employee_Hours_Historical_Current.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()

# --- Weekly upload ------------------------------------------------------------
new_file = st.file_uploader(
    "This week's Employee Hours export",
    type=["csv", "xlsx", "xlsm"],
    key="new_week_file",
)

if not new_file:
    st.info("Upload this week's Employee Hours export when you're ready.")
    st.stop()

file_bytes = new_file.getvalue()
file_hash = file_sha256(file_bytes)

try:
    new_df, new_meta = load_table_cached(file_bytes, new_file.name)
except Exception as exc:
    st.error(f"Could not read the weekly export: {exc}")
    st.stop()

period = normalize_period_display(new_meta.get("Time Period", ""))
if not period:
    st.error(
        "This workflow requires the weekly export's **Time Period** metadata so the "
        "app can prevent duplicate weeks and safely replace prior inputs."
    )
    st.stop()

period_key = canonical_period_key(period)
st.info(f"Weekly export period: **{period}**")

log_df = read_update_log(backend["log_ws"])
existing = find_log_row(log_df, period_key)

processed_now = False
replaced_now = False
merged = None
out_df = historical_df
replay_ambiguous = pd.DataFrame()
commit_record = None


def remember_result(period, file_hash, out_df, merged, replay_ambiguous):
    """Survive the rerun that a download-button click triggers."""
    st.session_state["last_result"] = {
        "period": period,
        "hash": file_hash,
        "out_df": out_df,
        "merged": merged,
        "replay_ambiguous": replay_ambiguous,
    }


if existing is None:
    # NEW PERIOD: append straight onto Historical. No replay needed.
    sequence = next_sequence(log_df)
    try:
        with st.spinner("Adding this week to Google Sheets..."):
            out_df, merged, replay_ambiguous, commit_record = process_week_into_google(
                backend, new_df, period, period_key, file_hash,
                new_file.name, sequence, mode="append",
            )
        processed_now = True
        remember_result(period, file_hash, out_df, merged, replay_ambiguous)
        st.success(f"✅ **{period}** was added automatically to the historical database.")
    except Exception as exc:
        st.error(f"The week could not be committed to Google Sheets: {exc}")
        st.stop()

else:
    existing_hash = str(existing.get("File Hash", ""))
    existing_status = str(existing.get("Status", "")).upper()
    sequence = int(pd.to_numeric(existing.get("Sequence", 0), errors="coerce") or 0)

    if existing_hash == file_hash and existing_status == "COMMITTED":
        st.warning(
            f"⚠️ **{period} is already inputted.** This is the exact same file, "
            "so no changes were made."
        )
        # Restore the results from the run that committed it, so they do not
        # vanish when a download click reruns the script.
        cached = st.session_state.get("last_result")
        if cached and cached["hash"] == file_hash:
            out_df = cached["out_df"]
            merged = cached["merged"]
            replay_ambiguous = cached["replay_ambiguous"]
            processed_now = True

    elif existing_hash == file_hash and existing_status != "COMMITTED":
        # A previous attempt was interrupted. Replay rather than append: the
        # interrupted run may have written a partial Historical, and appending
        # onto that would double-count.
        try:
            with st.spinner("Finishing the previously interrupted update..."):
                out_df, merged, replay_ambiguous, commit_record = process_week_into_google(
                    backend, new_df, period, period_key, file_hash,
                    new_file.name, sequence, mode="replay",
                )
            processed_now = True
            remember_result(period, file_hash, out_df, merged, replay_ambiguous)
            st.success(f"✅ **{period}** was completed successfully.")
        except Exception as exc:
            st.error(f"The interrupted week could not be completed: {exc}")
            st.stop()

    else:
        st.warning(
            f"⚠️ **{period} is already inputted, but this is a different file.**  \n\n"
            "Do you want to replace your past input with the file you are uploading now? "
            "If you replace it, the app will rebuild Historical from the Baseline and "
            "replay every stored week in order."
        )

        replace_col, keep_col = st.columns(2)
        with replace_col:
            replace_clicked = st.button(
                "Replace previous week with this file",
                type="primary",
                use_container_width=True,
            )
        with keep_col:
            keep_clicked = st.button("Keep existing week", use_container_width=True)

        if replace_clicked:
            try:
                with st.spinner("Replacing that week and rebuilding all later historical data..."):
                    out_df, merged, replay_ambiguous, commit_record = process_week_into_google(
                        backend, new_df, period, period_key, file_hash,
                        new_file.name, sequence, mode="replay",
                    )
                replaced_now = True
                remember_result(period, file_hash, out_df, merged, replay_ambiguous)
                st.success(
                    f"✅ **{period}** was replaced. Historical data was rebuilt from the "
                    "stored Baseline and weekly archive."
                )
            except Exception as exc:
                st.error(f"The prior week could not be replaced: {exc}")
                st.stop()
        elif keep_clicked:
            st.info(f"No changes were made. The existing **{period}** input was kept.")

# --- Results ------------------------------------------------------------------
if merged is not None:
    terminated = count_people(
        [merged["field_log"], merged["sync_log"]], "Active", "Terminated"
    )
    reactivated = count_people(
        [merged["field_log"], merged["sync_log"]], "Terminated", "Active"
    )

    st.subheader("This week's results")
    c1, c2 = st.columns(2)
    c1.metric("New hires added", f"{len(merged['new_hires']):,}")
    c2.metric("Active to Terminated", f"{terminated:,}")

    if reactivated:
        st.caption(f"{reactivated:,} went the other way, Terminated to Active.")

    # Kept because these silently lose data if ignored.
    if len(merged["dropped"]):
        lost = merged["dropped"]["Actual Hours"].sum(skipna=True)
        names = ", ".join(merged["dropped"]["Employee Full Name"].astype(str).head(5))
        st.warning(
            f"{len(merged['dropped'])} removed for not appearing in the export, "
            f"taking {lost:,.2f} cumulative hours: {names}"
            + (" ..." if len(merged["dropped"]) > 5 else "")
        )

    if len(merged["ambiguous"]):
        st.error(
            f"{len(merged['ambiguous'])} name(s) have several possible rehire matches "
            "and were left alone: "
            + ", ".join(merged["ambiguous"]["Employee Full Name"].astype(str))
        )

    # Replacing an older week replays later weeks too. Surface any ambiguous
    # matches found anywhere during that replay so none are silently hidden.
    if replaced_now and len(replay_ambiguous):
        later = replay_ambiguous[replay_ambiguous["Period"].astype(str) != period]
        if len(later):
            st.error(
                "The rebuild found ambiguous rehire match(es) in later stored week(s). "
                "Historical was rebuilt using the same existing rule (ambiguous matches "
                "are left alone). Review these rows:"
            )
            st.dataframe(later, use_container_width=True, hide_index=True)

# Always offer the freshest Historical after processing/replacement.
if processed_now or replaced_now:
    st.subheader("Updated historical file")
    label = re.sub(r"[^0-9A-Za-z]+", "_", period).strip("_")
    filename = f"Employee_Hours_Historical_{label}.xlsx"
    st.download_button(
        "Download updated historical file",
        data=build_excel_cached(out_df),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
